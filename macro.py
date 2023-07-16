import os
try:
    import openpyxl
    import shutil
    import getpass
    import time
    import pandas as pd
    import selenium
    import urllib
    import requests
    import webbrowser
    from selenium import webdriver
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    import win32com.client as win32
except:
    os.system("pip install --upgrade pip")
    os.system("pip install pip")
    os.system("pip install openpyxl")
    os.system("pip install pandas")
    os.system("pip install selenium")
    os.system("pip install urllib")
    os.system("pip install requests")
    os.system("pip install webbrowser")
    os.system("pip install win32com")
    os.system("pip install shutil")
    os.system("pip install getpass")
    import openpyxl
    import getpass
    import shutil
    import time
    import pandas as pd
    import selenium
    import urllib
    import requests
    import webbrowser
    from selenium import webdriver
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys

excel = win32.gencache.EnsureDispatch('Excel.Application')
for wb in excel.Workbooks:
    if wb.Name == "dados.xlsm":
        wb.Close(SaveChanges=True)
        break
excel.Quit()
time.sleep(3)
# carregando a planilha "Macro.xlsm"
try:
    wb = openpyxl.load_workbook("dados.xlsm", data_only=True)
    ws = wb.active
except Exception as e:
    print(e)
    input() 
# criando a lista única com as informações das colunas "A", "C", "E" e "F" para cada linha
lista_unica = [(ws.cell(row=row, column=1).value, ws.cell(row=row, column=3).value,
                ws.cell(row=row, column=5).value, ws.cell(row=row, column=6).value)
               for row in range(2, ws.max_row + 1)]
username = getpass.getuser()
options = webdriver.ChromeOptions()
options.add_argument(f"--user-data-dir=C:\\Users\\{username}\\AppData\\Local\\Google\\Chrome")
# Inicialize o driver
navegador = webdriver.Chrome(options=options)
# link = f"https://web.whatsapp.com/"
# navegador.get(link)
# time.sleep(10)
# WebDriverWait(navegador, 99999).until(EC.invisibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/div[1]/div/div/div[1]/ol/li[4]')))
for dados in lista_unica:
    # já estamos com o login feito no whatsapp web
    numero = dados[1]
    if (dados[1] == None) or (dados[2] == None):
        continue
    if len(numero) != 13:
        continue
    texto = urllib.parse.quote(dados[2])
    link = f"https://web.whatsapp.com/send?phone={numero}&text={texto}"
    time.sleep(1)
    navegador.get(link)
    time.sleep(1)
    # print("veri 1")
    # wait = WebDriverWait(navegador, 10)
    # enviar = wait.until(EC.presence_of_element_located((By.XPATH, "//button[@data-testid='compose-btn-send']")))
    # WebDriverWait(navegador, 10).until(EC.invisibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]')))
    cont = 0
    while True:
        # print("veri 2")
        cont += 1
        try:
            # print("veri 3")
            enviar = navegador.find_element(By.XPATH, "//button[@data-testid='compose-btn-send']")
            break
        except:
            try:
                # print("veri 4")
                enviar = navegador.find_element(By.XPATH, "/html/body/div[1]/div/span[2]/div/span/div/div/div/div/div/div[2]/div/div")
                enviar.click()
                cont = 999
                break
            except:
                # print("veri 5")
                time.sleep(3)
                if cont >= 20:
                    break
    if cont >= 20:
        continue
    # print("veri 6")
    enviar.click()
    if dados[3] != None:
        time.sleep(1)
        try:
            # Clica no botão adicionar
            arquivo = navegador.find_element(By.CSS_SELECTOR, "span[data-icon='clip']")
            arquivo.click()
            # Seleciona input
            time.sleep(1)
            attach = navegador.find_element(By.CSS_SELECTOR, "input[type='file']")
            # Adiciona arquivo
            attach.send_keys(dados[3])
            time.sleep(1)
            # Seleciona botão enviar
            send = navegador.find_element(By.XPATH,"/html/body/div[1]/div/div/div[2]/div[2]/span/div/span/div/div/div[2]/div/div[2]/div[2]/div/div")
            # Clica no botão enviar
            send.click()
        except:
            print("Erro ao enviar media")
            continue
time.sleep(10)    
print("concluido")
navegador.close()
